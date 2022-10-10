from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

#   load workbook
wb = load_workbook('algorithm_results.xlsx')

#   create sheet
sheet1 = wb.create_sheet('sheet1', 0)

#   activate worksheet to write dataframe
active = wb['sheet1']

#   Write dataframe to active worksheet
for x in dataframe_to_rows(df):
    active.append(x)

# Save workbook to write
wb.save(filepath)